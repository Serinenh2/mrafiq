"""Exports : registre Excel (§14/§39) et rapport de diagnostic PDF (§23)."""
import io
import os
from datetime import date
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable)
from .services import company_score, validation_report

INK = colors.HexColor('#0B2545'); PRIMARY = colors.HexColor('#1B5CB4')
BRASS = colors.HexColor('#B08A3C')

def _logo_path(company):
    logo = getattr(company, 'logo', None) if company else None
    if not logo:
        return None
    try:
        path = logo.path
        return path if os.path.exists(path) else None
    except Exception:
        return None

def company_logo_decorator(company):
    """Callback reportlab dessinant le logo de l'entreprise en haut de chaque page (import logo)."""
    path = _logo_path(company)
    if not path:
        return None
    try:
        reader = ImageReader(path)
        iw, ih = reader.getSize()
    except Exception:
        return None
    max_w, max_h = 26 * mm, 12 * mm
    ratio = min(max_w / iw, max_h / ih)
    w, h = iw * ratio, ih * ratio
    def _draw(canvas, doc):
        canvas.saveState()
        x = doc.pagesize[0] - doc.rightMargin - w
        y = doc.pagesize[1] - 4 * mm - h
        canvas.drawImage(reader, x, y, width=w, height=h, mask='auto', preserveAspectRatio=True)
        canvas.restoreState()
    return _draw

_noop_page = lambda canvas, doc: None

def build_pdf(doc, story, company):
    deco = company_logo_decorator(company) or _noop_page
    doc.build(story, onFirstPage=deco, onLaterPages=deco)

def add_logo_to_sheet(ws, company, anchor='A1', max_w=140, max_h=48):
    """Ajoute le logo de l'entreprise (image flottante) en haut de la feuille Excel."""
    path = _logo_path(company)
    if not path:
        return
    try:
        iw, ih = PILImage.open(path).size
        ratio = min(max_w / iw, max_h / ih, 1)
        img = XlsxImage(path)
        img.width, img.height = iw * ratio, ih * ratio
        img.anchor = anchor
        ws.add_image(img)
    except Exception:
        pass

def registre_xlsx(company):
    wb = Workbook(); ws = wb.active; ws.title = 'Registre'
    headers = ['Réf.', 'Traitement', 'Responsable', 'Service', 'Finalité', 'Personnes',
               'Données', 'Destinataires', 'Conservation', 'Sous-traitants',
               'Transfert étranger', 'Statut', 'Conformité (%)']
    ws.append(headers)
    fill = PatternFill('solid', fgColor='0B2545')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
        c.alignment = Alignment(vertical='center')
    for p in company.processings.exclude(status='rejete'):
        vals = {'conforme':100,'partiel':50,'non_conforme':0,'manquant':0}
        pts = [vals[a.status] for a in p.assessments.all() if a.status in vals]
        ws.append([
            p.reference, p.name, p.owner_name, p.department, p.purpose,
            ', '.join(s.label_fr for s in p.subject_categories.all()),
            ', '.join(d.label_fr for d in p.data_categories.all()),
            p.recipients,
            p.retention_duration,
            ', '.join(pr.name for pr in p.processors.all()),
            'Oui' if p.transfer_abroad else 'Non',
            p.get_status_display(),
            round(sum(pts)/len(pts)) if pts else '',
        ])
    for i, w in enumerate([10,26,18,16,34,24,28,24,16,22,14,14,12], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.insert_rows(1, amount=2); ws.row_dimensions[1].height = 36
    add_logo_to_sheet(ws, company)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="registre_{company.pk}.xlsx"'
    return resp

def diagnostic_pdf(company):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm)
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=ss['Title'], textColor=INK, fontSize=20, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=ss['Normal'], textColor=BRASS, fontSize=9, spaceAfter=14)
    h2 = ParagraphStyle('h2', parent=ss['Heading2'], textColor=PRIMARY, fontSize=13, spaceBefore=14)
    body = ss['BodyText']
    score = company_score(company)
    story = [
        Paragraph('MRAFIQ — Rapport de diagnostic de conformité', h1),
        Paragraph(f'{company.name} · {date.today().strftime("%d/%m/%Y")} · Loi n° 18-07', sub),
        HRFlowable(width='100%', color=BRASS, thickness=2), Spacer(1, 6),
        Paragraph('1. Synthèse exécutive', h2),
    ]
    if score['global'] is not None:
        story.append(Paragraph(
            f"Score global de conformité : <b>{score['global']} %</b> — niveau "
            f"<b>{score['level']['fr']}</b>. Ce score agrège l'évaluation des exigences "
            f"configurées dans le référentiel juridique, pondérées par domaine.", body))
    else:
        story.append(Paragraph("Aucune évaluation disponible : lancer le moteur de conformité.", body))

    story.append(Paragraph('2. Score par domaine', h2))
    rows = [['Domaine', 'Poids', 'Score']]
    for d in score['domains']:
        rows.append([d['label_fr'], str(d['weight']),
                     f"{d['score']} %" if d['score'] is not None else '—'])
    t = Table(rows, colWidths=[95*mm, 25*mm, 30*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), INK), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), .4, colors.HexColor('#C6D1E0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F7FB')]),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5)]))
    story.append(t)

    story.append(Paragraph('3. Principaux écarts', h2))
    rows = [['Réf. traitement', 'Exigence', 'Gravité', 'Description']]
    for g in company.gaps.filter(is_open=True).select_related('processing','requirement')[:20]:
        rows.append([g.processing.reference if g.processing else '—',
                     g.requirement.code if g.requirement else '—',
                     g.get_severity_display(), Paragraph(g.description, ss['BodyText'])])
    if len(rows) == 1: rows.append(['—','—','—','Aucun écart ouvert.'])
    t = Table(rows, colWidths=[26*mm, 22*mm, 20*mm, 82*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8.5), ('GRID', (0,0), (-1,-1), .4, colors.HexColor('#C6D1E0')),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
    story.append(t)

    story.append(Paragraph('4. Actions prioritaires', h2))
    for a in company.actions.exclude(status__in=['termine','valide']).order_by('due_date')[:10]:
        story.append(Paragraph(
            f"<b>{a.reference}</b> — {a.title} · priorité {a.get_priority_display() if hasattr(a,'get_priority_display') else a.priority}"
            f"{' · échéance ' + a.due_date.strftime('%d/%m/%Y') if a.due_date else ''}", body))

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Document préparé via la plateforme MRAFIQ à des fins d'accompagnement. "
        "Il ne constitue ni une décision ni une autorisation d'une autorité administrative.",
        ParagraphStyle('foot', parent=ss['Normal'], fontSize=8, textColor=colors.HexColor('#51617A'))))
    build_pdf(doc, story, company)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="rapport_diagnostic_{company.pk}.pdf"'
    return resp

def declaration_pdf(company):
    """Dossier de déclaration (§21-23) : identification, responsable/DPO, traitements, sécurité, droits."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm)
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=ss['Title'], textColor=INK, fontSize=20, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=ss['Normal'], textColor=BRASS, fontSize=9, spaceAfter=14)
    h2 = ParagraphStyle('h2', parent=ss['Heading2'], textColor=PRIMARY, fontSize=13, spaceBefore=14)
    body = ss['BodyText']
    report = validation_report(company)

    story = [
        Paragraph('MRAFIQ — Dossier de déclaration', h1),
        Paragraph(f'{company.name} · {date.today().strftime("%d/%m/%Y")} · Loi n° 18-07', sub),
        HRFlowable(width='100%', color=BRASS, thickness=2), Spacer(1, 6),
        Paragraph("1. Identification de l'entreprise", h2),
    ]
    rows = [
        ['Raison sociale', company.name], ['Forme juridique', company.legal_form or '—'],
        ['Registre de commerce', company.rc_number or '—'],
        ['NIF', company.nif or '—'], ['NIS', company.nis or '—'],
        ['Adresse', f'{company.address or "—"}, {company.commune or ""} {company.wilaya or ""}'.strip(', ')],
        ['Secteur', company.sector or '—'], ['Effectif', str(company.employees_count)],
    ]
    t = Table(rows, colWidths=[50*mm, 125*mm])
    t.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), .4, colors.HexColor('#C6D1E0')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F5F7FB')),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
    story.append(t)

    story.append(Paragraph('2. Responsable des traitements et DPO', h2))
    story.append(Paragraph(
        f"<b>Responsable des traitements :</b> {company.controller_name or 'Non renseigné'} "
        f"({company.controller_qualite or '—'}) · {company.controller_phone or '—'} · "
        f"{company.controller_email or '—'}", body))
    dpo_labels = {'aucun': 'Sans DPO désigné', 'en_cours': 'Désignation en cours', 'designe': 'Désigné'}
    dpo_line = f"<b>DPO :</b> {dpo_labels[company.dpo_status]}"
    if company.dpo_status != 'aucun':
        dpo_line += f" — {company.dpo_name or '—'} · {company.dpo_phone or '—'} · {company.dpo_email or '—'}"
    story.append(Paragraph(dpo_line, body))

    story.append(Paragraph('3. Traitements déclarés', h2))
    rows = [['Réf.', 'Traitement', 'Finalité', 'Statut']]
    for p in company.processings.exclude(status__in=['propose', 'rejete']):
        rows.append([p.reference, p.name, Paragraph((p.purpose or '—')[:150], ss['BodyText']),
                     p.get_status_display()])
    if len(rows) == 1:
        rows.append(['—', 'Aucun traitement', '—', '—'])
    t = Table(rows, colWidths=[20*mm, 38*mm, 92*mm, 25*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), PRIMARY), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8.5), ('GRID', (0,0), (-1,-1), .4, colors.HexColor('#C6D1E0')),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4)]))
    story.append(t)

    story.append(Paragraph('4. Sécurité et droits des personnes', h2))
    story.append(Paragraph(
        f"Mesures de sécurité en place : <b>{report['security_total'] - report['security_todo']} "
        f"/ {report['security_total']}</b>. Droits des personnes à vérifier : "
        f"<b>{report['rights_todo']} / 4</b>.", body))

    story.append(Paragraph('5. Écarts et actions', h2))
    story.append(Paragraph(
        f"Écarts ouverts : <b>{report['open_gaps']}</b>, dont critiques : "
        f"<b>{report['critical_gaps']}</b>.", body))

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Ce dossier a été généré via la plateforme MRAFIQ à partir des informations saisies par le "
        "consultant. Il constitue un outil d'aide à la déclaration et ne dispense pas d'une "
        "vérification juridique finale.",
        ParagraphStyle('foot', parent=ss['Normal'], fontSize=8, textColor=colors.HexColor('#51617A'))))
    build_pdf(doc, story, company)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="declaration_{company.pk}.pdf"'
    return resp

def company_profile_pdf(company):
    """Fiche organisation seule (§20)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18*mm, bottomMargin=18*mm)
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=ss['Title'], textColor=INK, fontSize=20, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=ss['Normal'], textColor=BRASS, fontSize=9, spaceAfter=14)
    body = ss['BodyText']
    story = [
        Paragraph('MRAFIQ — Fiche entreprise', h1),
        Paragraph(f'{date.today().strftime("%d/%m/%Y")} · Loi n° 18-07', sub),
        HRFlowable(width='100%', color=BRASS, thickness=2), Spacer(1, 10),
    ]
    rows = [
        ['Raison sociale', company.name], ['Forme juridique', company.legal_form or '—'],
        ['Secteur', company.sector or '—'], ['Activité principale', company.main_activity or '—'],
        ['Registre de commerce', company.rc_number or '—'],
        ['NIF', company.nif or '—'], ['NIS', company.nis or '—'],
        ['Adresse', company.address or '—'],
        ['Wilaya', company.wilaya or '—'], ['Commune', company.commune or '—'],
        ['Site internet', company.website or '—'], ['Effectif', str(company.employees_count)],
        ['Contact', f'{company.contact_name or "—"} · {company.contact_email or "—"} · {company.contact_phone or "—"}'],
        ['Responsable des traitements', f'{company.controller_name or "—"} ({company.controller_qualite or "—"})'],
        ['DPO', {'aucun': 'Sans DPO', 'en_cours': 'Désignation en cours', 'designe': 'Désigné'}[company.dpo_status]
                + (f' — {company.dpo_name}' if company.dpo_status != 'aucun' and company.dpo_name else '')],
        ['Systèmes informatiques', company.it_systems or '—'],
        ['Prestataires informatiques', company.it_providers or '—'],
        ['Organisation interne', company.internal_organisation or '—'],
    ]
    t = Table(rows, colWidths=[55*mm, 120*mm])
    t.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 9), ('GRID', (0,0), (-1,-1), .4, colors.HexColor('#C6D1E0')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F5F7FB')), ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5)]))
    story.append(t)
    build_pdf(doc, story, company)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="fiche_entreprise_{company.pk}.pdf"'
    return resp

def actions_xlsx(company):
    """Plan d'action (§17/§41)."""
    wb = Workbook(); ws = wb.active; ws.title = "Plan d'action"
    headers = ['Réf.', 'Titre', 'Domaine', 'Priorité', 'Responsable', 'Échéance', 'Statut', 'Traitement', 'Commentaire']
    ws.append(headers)
    fill = PatternFill('solid', fgColor='0B2545')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
        c.alignment = Alignment(vertical='center')
    for a in company.actions.select_related('processing', 'gap__requirement__domain').order_by('due_date'):
        domain = a.gap.requirement.domain.label_fr if a.gap and a.gap.requirement else ''
        ws.append([
            a.reference, a.title, domain, a.get_priority_display(), a.assignee,
            a.due_date.strftime('%d/%m/%Y') if a.due_date else '', a.get_status_display(),
            a.processing.reference if a.processing else '', a.comment,
        ])
    for i, w in enumerate([10, 40, 16, 12, 18, 14, 14, 12, 30], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.insert_rows(1, amount=2); ws.row_dimensions[1].height = 36
    add_logo_to_sheet(ws, company)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="plan_action_{company.pk}.xlsx"'
    return resp

def cartographie_xlsx(company):
    """Cartographie des flux (§19/§41) — table des relations personnes → traitement → destinataires."""
    wb = Workbook(); ws = wb.active; ws.title = 'Cartographie'
    headers = ['Traitement', 'Service', 'Personnes concernées', 'Catégories de données',
               'Sous-traitants', 'Destinataires', 'Transfert à l\'étranger']
    ws.append(headers)
    fill = PatternFill('solid', fgColor='0B2545')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = fill
        c.alignment = Alignment(vertical='center')
    for p in company.processings.exclude(status__in=['propose', 'rejete']):
        ws.append([
            f'{p.reference} · {p.name}', p.department,
            ', '.join(s.label_fr for s in p.subject_categories.all()),
            ', '.join(d.label_fr for d in p.data_categories.all()),
            ', '.join(pr.name for pr in p.processors.all()),
            p.recipients,
            (p.transfer_country or 'Oui') if p.transfer_abroad else 'Non',
        ])
    for i, w in enumerate([32, 16, 26, 26, 22, 26, 18], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.insert_rows(1, amount=2); ws.row_dimensions[1].height = 36
    add_logo_to_sheet(ws, company)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="cartographie_{company.pk}.xlsx"'
    return resp
